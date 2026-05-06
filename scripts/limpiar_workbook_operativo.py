"""Mantenimiento controlado del workbook operativo local.

El script crea un respaldo antes de modificar el archivo real y solo elimina la
hoja obsoleta configurada. No lee ni reporta datos de filas.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.core.exceptions import GestorSegurosError


OBSOLETE_SHEET_NAME = "Reporte de vencimientos del mes"
REPORT_MD_NAME = "reporte_limpieza_workbook.md"
REPORT_JSON_NAME = "reporte_limpieza_workbook.json"


class WorkbookMaintenanceError(GestorSegurosError):
    """Se usa cuando el mantenimiento del Control Cartera no puede completarse con seguridad."""


@dataclass(frozen=True)
class MaintenanceResult:
    """Resultado del mantenimiento controlado del Control Cartera."""

    input_path: Path
    backup_path: Path
    markdown_report: Path
    json_report: Path
    sheet_found: bool
    sheet_deleted: bool


def clean_operational_workbook(
    input_path: str | Path,
    backup_dir: str | Path,
    report_dir: str | Path,
    obsolete_sheet_name: str = OBSOLETE_SHEET_NAME,
) -> MaintenanceResult:
    """Elimina una hoja obsoleta del Control Cartera después de crear respaldo."""
    workbook_path = Path(input_path).resolve()
    backups_path = Path(backup_dir).resolve()
    reports_path = Path(report_dir).resolve()

    _validate_input(workbook_path)
    backups_path.mkdir(parents=True, exist_ok=True)
    reports_path.mkdir(parents=True, exist_ok=True)

    generated_at = datetime.now()
    backup_path = _create_backup(workbook_path, backups_path, generated_at)

    workbook = load_workbook(workbook_path)
    try:
        sheets_before = list(workbook.sheetnames)
        sheet_found = obsolete_sheet_name in workbook.sheetnames
        sheet_deleted = False
        warnings: list[str] = []

        if sheet_found:
            if len(workbook.sheetnames) <= 1:
                raise WorkbookMaintenanceError("No se puede eliminar la unica hoja del workbook.")
            del workbook[obsolete_sheet_name]
            workbook.save(workbook_path)
            sheet_deleted = True
        else:
            warnings.append("La hoja obsoleta no fue encontrada; el workbook no fue modificado.")

        sheets_after = list(workbook.sheetnames)
    finally:
        workbook.close()

    report = _build_report(
        generated_at=generated_at,
        input_path=workbook_path,
        backup_path=backup_path,
        obsolete_sheet_name=obsolete_sheet_name,
        sheet_found=sheet_found,
        sheet_deleted=sheet_deleted,
        sheets_before=sheets_before,
        sheets_after=sheets_after,
        warnings=warnings,
    )

    markdown_report = reports_path / REPORT_MD_NAME
    json_report = reports_path / REPORT_JSON_NAME
    markdown_report.write_text(_render_markdown_report(report), encoding="utf-8")
    json_report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    return MaintenanceResult(
        input_path=workbook_path,
        backup_path=backup_path,
        markdown_report=markdown_report,
        json_report=json_report,
        sheet_found=sheet_found,
        sheet_deleted=sheet_deleted,
    )


def _validate_input(workbook_path: Path) -> None:
    if not workbook_path.exists():
        raise WorkbookMaintenanceError(f"No existe el workbook de entrada: {workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise WorkbookMaintenanceError("El workbook de entrada debe ser .xlsx o .xlsm.")


def _create_backup(workbook_path: Path, backup_dir: Path, generated_at: datetime) -> Path:
    timestamp = generated_at.strftime("%Y%m%d_%H%M%S")
    backup_name = f"{workbook_path.stem}_backup_limpieza_{timestamp}{workbook_path.suffix}"
    backup_path = backup_dir / backup_name
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def _build_report(
    generated_at: datetime,
    input_path: Path,
    backup_path: Path,
    obsolete_sheet_name: str,
    sheet_found: bool,
    sheet_deleted: bool,
    sheets_before: list[str],
    sheets_after: list[str],
    warnings: list[str],
) -> dict[str, Any]:
    return {
        "generated_at": generated_at.isoformat(timespec="seconds"),
        "input_path": str(input_path),
        "backup_path": str(backup_path),
        "obsolete_sheet_name": obsolete_sheet_name,
        "sheet_found": sheet_found,
        "sheet_deleted": sheet_deleted,
        "sheets_before": sheets_before,
        "sheets_after": sheets_after,
        "warnings": warnings,
        "next_steps": [
            "Validar manualmente que la hoja principal de cartera conserva sus datos.",
            "Mantener el respaldo local generado mientras se revisa el workbook limpio.",
            "Implementar vencimientos mensuales dentro del sistema en una fase futura.",
        ],
        "privacy": {
            "contains_row_data": False,
            "notes": "El reporte contiene nombres de hojas y rutas locales, no datos de filas.",
        },
    }


def _render_markdown_report(report: dict[str, Any]) -> str:
    return "\n".join(
        [
            "# Reporte de Limpieza Workbook",
            "",
            "Reporte local de mantenimiento. No incluye datos de filas.",
            "",
            f"- Generado: `{report['generated_at']}`",
            f"- Workbook: `{report['input_path']}`",
            f"- Respaldo: `{report['backup_path']}`",
            f"- Hoja obsoleta: `{report['obsolete_sheet_name']}`",
            f"- Hoja encontrada: `{report['sheet_found']}`",
            f"- Hoja eliminada: `{report['sheet_deleted']}`",
            "",
            "## Hojas Antes",
            "",
            _format_bullets(report["sheets_before"]),
            "",
            "## Hojas Despues",
            "",
            _format_bullets(report["sheets_after"]),
            "",
            "## Advertencias",
            "",
            _format_bullets(report["warnings"]),
            "",
            "## Proximos Pasos",
            "",
            _format_bullets(report["next_steps"]),
            "",
        ]
    )


def _format_bullets(items: list[str]) -> str:
    if not items:
        return "- Sin registros."
    return "\n".join(f"- {item}" for item in items)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Limpia una hoja obsoleta del workbook operativo.")
    parser.add_argument("input_path", help="Ruta del workbook operativo real.")
    parser.add_argument("backup_dir", help="Carpeta local para respaldos.")
    parser.add_argument("report_dir", help="Carpeta local para reportes de mantenimiento.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        result = clean_operational_workbook(args.input_path, args.backup_dir, args.report_dir)
    except GestorSegurosError as exc:
        print(f"Error de mantenimiento: {exc}")
        return 1

    print("Mantenimiento de workbook completado.")
    print(f"- Respaldo: {result.backup_path}")
    print(f"- Reporte Markdown: {result.markdown_report}")
    print(f"- Reporte JSON: {result.json_report}")
    print(f"- Hoja eliminada: {result.sheet_deleted}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
