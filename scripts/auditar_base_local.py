"""Auditoria segura de la base local confidencial.

El script analiza estructura, completitud y patrones generales de un workbook
Excel sin modificar el archivo original. Los reportes generados no incluyen
muestras de filas ni valores reales potencialmente sensibles.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FREQUENCY_DM = "D.M. / deduccion mensual"
FREQUENCY_MONTHLY = "mensual"
FREQUENCY_QUARTERLY = "trimestral"
FREQUENCY_SEMIANNUAL = "semestral"
FREQUENCY_ANNUAL = "anual"
FREQUENCY_OTHER = "otros formatos observados"
FREQUENCY_EMPTY = "vacio"

POLICY_NUMERIC = "completamente numerico"
POLICY_PREFIX_01 = "prefijo_01_colones_preliminar"
POLICY_PREFIX_02 = "prefijo_02_dolares_preliminar"
POLICY_OTHER = "otros_formatos"
POLICY_EMPTY = "vacio"

ID_PHYSICAL = "cedula_fisica_probable"
ID_LEGAL_OR_NUMERIC = "cedula_juridica_u_otro_numerico_probable"
ID_PASSPORT_OR_FOREIGN = "pasaporte_o_extranjero_probable"
ID_OTHER_NUMERIC = "otro_formato_numerico"
ID_OTHER = "otros_formatos"
ID_EMPTY = "vacio"

HEADER_KEYWORDS = (
    "asegurado",
    "aseguradora",
    "cliente",
    "correo",
    "detalle",
    "direccion",
    "fecha",
    "finca",
    "frecuencia",
    "identificacion",
    "mes",
    "moneda",
    "nombre",
    "pasaporte",
    "placa",
    "poliza",
    "prima",
    "telefono",
    "venc",
    "vigencia",
    "ano",
    "anio",
    "dia",
)
HEADER_CONFIDENCE_MIN_KEYWORDS = 2
HEADER_CONFIDENCE_MIN_SCORE = 18


def is_empty(value: Any) -> bool:
    """Devuelve True para celdas vacias o texto sin contenido."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def safe_cell_text(value: Any) -> str:
    """Convierte una celda a texto solo para clasificacion interna."""
    if is_empty(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_text(value: Any) -> str:
    """Normaliza texto para busquedas flexibles sin alterar datos de origen."""
    text = safe_cell_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\u00ba", "o").replace("\u00b0", "o")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def technical_column_id(index: int) -> str:
    """Identificador técnico estable para reportar columnas sin PII."""
    return f"COL_{get_column_letter(index + 1)}"


def classify_frequency(value: Any) -> str:
    """Clasifica vigencias/frecuencias en categorias observadas no sensibles."""
    normalized = normalize_text(value)
    compact = normalized.replace(" ", "")

    if not normalized:
        return FREQUENCY_EMPTY
    if compact in {"dm", "dms"} or "deduccion mensual" in normalized:
        return FREQUENCY_DM
    if "trimes" in normalized or compact in {"trim", "trimestral"}:
        return FREQUENCY_QUARTERLY
    if "semes" in normalized or compact in {"semi", "semestral"}:
        return FREQUENCY_SEMIANNUAL
    if "mensual" in normalized or compact in {"mens", "mensuales"}:
        return FREQUENCY_MONTHLY
    if "anual" in normalized or compact in {"anuales", "annual"}:
        return FREQUENCY_ANNUAL
    return FREQUENCY_OTHER


def summarize_frequency_values(values: Iterable[Any]) -> dict[str, Any]:
    """Resume categorias de vigencia/frecuencia sin listar valores originales."""
    counter = Counter(classify_frequency(value) for value in values)
    return {
        "has_dm": counter.get(FREQUENCY_DM, 0) > 0,
        "categories": dict(sorted(counter.items())),
    }


def classify_policy_number(value: Any) -> str:
    """Clasifica números de póliza sin exponer el número completo."""
    text = safe_cell_text(value)
    compact = re.sub(r"\s+", "", text)

    if not compact:
        return POLICY_EMPTY
    if compact.isdigit():
        return POLICY_NUMERIC
    if compact.startswith("01"):
        return POLICY_PREFIX_01
    if compact.startswith("02"):
        return POLICY_PREFIX_02
    return POLICY_OTHER


def classify_identification_format(value: Any) -> str:
    """Clasificación conservadora de formatos de identificación."""
    text = safe_cell_text(value)
    compact = re.sub(r"[^A-Za-z0-9]+", "", text).upper()

    if not compact:
        return ID_EMPTY
    if any(ch.isalpha() for ch in compact):
        return ID_PASSPORT_OR_FOREIGN
    if compact.isdigit() and len(compact) == 9:
        return ID_PHYSICAL
    if compact.isdigit() and len(compact) >= 10:
        return ID_LEGAL_OR_NUMERIC
    if compact.isdigit():
        return ID_OTHER_NUMERIC
    return ID_OTHER


def count_header_keywords(value: Any) -> int:
    normalized = normalize_text(value)
    return sum(1 for keyword in HEADER_KEYWORDS if keyword in normalized.split() or keyword in normalized)


def looks_like_sensitive_data(value: Any) -> bool:
    """Detecta valores que no deben tratarse como encabezados reportables."""
    if is_empty(value):
        return False
    if isinstance(value, (datetime, date)):
        return True

    text = safe_cell_text(value)
    normalized = normalize_text(text)
    compact = re.sub(r"[^A-Za-z0-9]+", "", text)
    digits = re.sub(r"\D+", "", text)

    if "@" in text and "." in text:
        return True
    if text.startswith(("01", "02")) and any(ch.isdigit() for ch in text):
        return True
    if compact.isdigit() and len(compact) >= 4:
        return True
    if len(digits) >= 6:
        return True
    if classify_frequency(text) not in {FREQUENCY_EMPTY, FREQUENCY_OTHER}:
        return True
    if re.fullmatch(r"[A-Z]{2,4}\d{2,5}", compact.upper()):
        return True
    if len(normalized.split()) >= 4 and count_header_keywords(text) == 0:
        return True
    return False


def is_safe_header_label(value: Any) -> bool:
    """Confirma si una celda puede mostrarse como nombre de columna."""
    text = safe_cell_text(value)
    if not text or len(text) > 80:
        return False
    if not re.search(r"[A-Za-z]", text):
        return False
    if looks_like_sensitive_data(text):
        return False
    return True


def evaluate_header_candidate(row: Iterable[Any]) -> dict[str, Any]:
    """Puntua una fila candidata sin almacenar valores de datos reales."""
    values = list(row)
    non_empty = [value for value in values if not is_empty(value)]
    keyword_matches = sum(1 for value in non_empty if count_header_keywords(value) > 0)
    safe_labels = sum(1 for value in non_empty if is_safe_header_label(value))
    sensitive_like = sum(1 for value in non_empty if looks_like_sensitive_data(value))
    score = keyword_matches * 10 + safe_labels * 2 + len(non_empty) - sensitive_like * 6
    confirmed = (
        keyword_matches >= HEADER_CONFIDENCE_MIN_KEYWORDS
        and score >= HEADER_CONFIDENCE_MIN_SCORE
        and safe_labels >= max(2, keyword_matches)
    )

    return {
        "score": score,
        "confirmed": confirmed,
        "keyword_matches": keyword_matches,
        "safe_labels": safe_labels,
        "sensitive_like": sensitive_like,
        "filled_count": len(non_empty),
    }


def detect_header_info(worksheet: Any, scan_rows: int = 20) -> dict[str, Any]:
    """Estima la fila de encabezados y exige confianza antes de reportarla."""
    max_scan_row = min(worksheet.max_row or 1, scan_rows)
    best = {
        "row": None,
        "candidate_row": None,
        "confirmed": False,
        "score": -1,
        "keyword_matches": 0,
        "safe_labels": 0,
        "sensitive_like": 0,
        "filled_count": 0,
    }

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True), start=1
    ):
        evaluation = evaluate_header_candidate(row)
        if evaluation["score"] > best["score"]:
            best.update(evaluation)
            best["candidate_row"] = row_index
            best["row"] = row_index if evaluation["confirmed"] else None

    if not best["confirmed"]:
        best["row"] = None
    return best


def get_row_values(worksheet: Any, row_number: int | None) -> list[Any]:
    if row_number is None:
        return []
    row = next(
        worksheet.iter_rows(
            min_row=row_number,
            max_row=row_number,
            max_col=worksheet.max_column or 1,
            values_only=True,
        ),
        (),
    )
    return list(row)


def build_column_metadata(
    raw_headers: list[Any], column_count: int, header_confirmed: bool
) -> list[dict[str, Any]]:
    """Crea metadatos seguros; nunca guarda encabezados dudosos."""
    columns = []
    for index in range(column_count):
        technical_id = technical_column_id(index)
        raw_header = raw_headers[index] if header_confirmed and index < len(raw_headers) else None
        confirmed_name = safe_cell_text(raw_header) if is_safe_header_label(raw_header) else None
        columns.append(
            {
                "index": index + 1,
                "letter": get_column_letter(index + 1),
                "technical_id": technical_id,
                "display_name": confirmed_name or technical_id,
                "confirmed_name": confirmed_name,
                "header_confirmed": bool(confirmed_name),
            }
        )
    return columns


def column_search_text(column: dict[str, Any]) -> str:
    return normalize_text(column.get("confirmed_name") or "")


def header_contains(column: dict[str, Any], *keywords: str) -> bool:
    normalized = column_search_text(column)
    return any(keyword in normalized for keyword in keywords)


def find_column_indexes(columns: list[dict[str, Any]], *keywords: str) -> list[int]:
    """Devuelve indices cero-basados de columnas confirmadas por encabezado."""
    return [index for index, column in enumerate(columns) if header_contains(column, *keywords)]


def detect_separate_due_date_columns(columns: list[dict[str, Any]]) -> dict[str, Any]:
    """Detecta campos separados de dia, mes y ano para vencimiento."""
    result = {"dia": [], "mes": [], "ano": []}

    for index, column in enumerate(columns):
        normalized = column_search_text(column)
        tokens = set(normalized.split())
        payload = {
            "index": index + 1,
            "technical_id": column["technical_id"],
            "display_name": column["display_name"],
        }

        if "dia" in tokens and ("venc" in normalized or normalized == "dia"):
            result["dia"].append(payload)
        if "mes" in tokens and ("venc" in normalized or normalized == "mes"):
            result["mes"].append(payload)
        if ("ano" in tokens or "anio" in tokens) and (
            "venc" in normalized or normalized in {"ano", "anio"}
        ):
            result["ano"].append(payload)

    return {
        "has_separate_due_date_fields": bool(result["dia"] and result["mes"] and result["ano"]),
        "columns": result,
    }


def summarize_sheet(worksheet: Any) -> dict[str, Any]:
    header_info = detect_header_info(worksheet)
    header_row = header_info["row"]
    raw_headers = get_row_values(worksheet, header_row)
    data_start_row = header_row + 1 if header_row else 1
    data_rows = 0

    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        if any(not is_empty(value) for value in row):
            data_rows += 1

    return {
        "name": worksheet.title,
        "max_row": worksheet.max_row or 0,
        "max_column": worksheet.max_column or 0,
        "header_row": header_row,
        "header_candidate_row": header_info["candidate_row"],
        "header_confirmed": bool(header_row),
        "header_confidence_score": header_info["score"],
        "header_keyword_matches": header_info["keyword_matches"],
        "header_count": sum(1 for header in raw_headers if is_safe_header_label(header)),
        "data_rows_estimated": data_rows,
    }


def select_main_sheet(sheet_summaries: list[dict[str, Any]]) -> dict[str, Any]:
    """Selecciona la hoja principal por volumen de datos y encabezados seguros."""
    return max(
        sheet_summaries,
        key=lambda item: (
            item["data_rows_estimated"],
            item["header_count"],
            item["max_column"],
        ),
    )


def is_critical_column(column: dict[str, Any]) -> bool:
    normalized = column_search_text(column)
    keywords = (
        "cliente",
        "nombre",
        "asegurado",
        "identificacion",
        "cedula",
        "pasaporte",
        "poliza",
        "vigencia",
        "frecuencia",
        "vencimiento",
        "fecha",
        "dia",
        "mes",
        "ano",
        "prima",
        "moneda",
    )
    return any(keyword in normalized for keyword in keywords)


def recognized_frequency_count(counter: Counter) -> int:
    return sum(
        counter.get(category, 0)
        for category in (
            FREQUENCY_DM,
            FREQUENCY_MONTHLY,
            FREQUENCY_QUARTERLY,
            FREQUENCY_SEMIANNUAL,
            FREQUENCY_ANNUAL,
        )
    )


def should_report_policy_counter(counter: Counter) -> bool:
    return (
        counter.get(POLICY_PREFIX_01, 0) > 0
        or counter.get(POLICY_PREFIX_02, 0) > 0
        or counter.get(POLICY_NUMERIC, 0) > 0
    )


def analyze_main_sheet(worksheet: Any, sheet_summary: dict[str, Any]) -> dict[str, Any]:
    header_row = sheet_summary["header_row"]
    raw_headers = get_row_values(worksheet, header_row)
    column_count = worksheet.max_column or len(raw_headers)
    columns = build_column_metadata(raw_headers, column_count, bool(header_row))
    data_start_row = header_row + 1 if header_row else 1

    column_stats = [
        {
            "index": column["index"],
            "letter": column["letter"],
            "technical_id": column["technical_id"],
            "display_name": column["display_name"],
            "confirmed_name": column["confirmed_name"],
            "header_confirmed": column["header_confirmed"],
            "filled_count": 0,
            "empty_count": 0,
            "empty_percentage": 0.0,
            "critical": is_critical_column(column),
        }
        for column in columns
    ]

    frequency_columns = find_column_indexes(
        columns, "vigencia", "frecuencia", "periodicidad", "forma pago", "modalidad pago"
    )
    policy_columns = find_column_indexes(columns, "poliza")
    identification_columns = find_column_indexes(columns, "identificacion", "cedula", "pasaporte")
    detail_columns = find_column_indexes(columns, "detalle")
    plate_farm_columns = find_column_indexes(columns, "placa", "finca")

    observed_frequency_by_column = {index: Counter() for index in range(column_count)}
    observed_policy_by_column = {index: Counter() for index in range(column_count)}
    identification_by_column = {index: Counter() for index in identification_columns}
    detail_counts = {index: {"filled_count": 0, "empty_count": 0} for index in detail_columns}
    plate_farm_counts = {index: {"filled_count": 0, "empty_count": 0} for index in plate_farm_columns}
    data_rows = 0

    for row in worksheet.iter_rows(min_row=data_start_row, max_col=column_count, values_only=True):
        if not any(not is_empty(value) for value in row):
            continue

        data_rows += 1
        for index in range(column_count):
            value = row[index] if index < len(row) else None
            if is_empty(value):
                column_stats[index]["empty_count"] += 1
            else:
                column_stats[index]["filled_count"] += 1

            observed_frequency_by_column[index][classify_frequency(value)] += 1
            observed_policy_by_column[index][classify_policy_number(value)] += 1

        for index in identification_columns:
            identification_by_column[index][
                classify_identification_format(row[index] if index < len(row) else None)
            ] += 1
        for index in detail_columns:
            bucket = "empty_count" if is_empty(row[index] if index < len(row) else None) else "filled_count"
            detail_counts[index][bucket] += 1
        for index in plate_farm_columns:
            bucket = "empty_count" if is_empty(row[index] if index < len(row) else None) else "filled_count"
            plate_farm_counts[index][bucket] += 1

    if not frequency_columns:
        frequency_columns = [
            index
            for index, counter in observed_frequency_by_column.items()
            if recognized_frequency_count(counter) > 0
        ]
    if not policy_columns:
        policy_columns = [
            index
            for index, counter in observed_policy_by_column.items()
            if should_report_policy_counter(counter)
        ]

    frequency_by_column = {
        index: observed_frequency_by_column[index] for index in frequency_columns
    }
    policy_by_column = {index: observed_policy_by_column[index] for index in policy_columns}

    for stat in column_stats:
        if data_rows:
            stat["empty_percentage"] = round((stat["empty_count"] / data_rows) * 100, 2)

    critical_issues = [
        {
            "index": stat["index"],
            "letter": stat["letter"],
            "technical_id": stat["technical_id"],
            "display_name": stat["display_name"],
            "empty_count": stat["empty_count"],
            "empty_percentage": stat["empty_percentage"],
        }
        for stat in column_stats
        if stat["critical"] and stat["empty_count"] > 0
    ]

    return {
        "dimensions": {
            "max_row": worksheet.max_row or 0,
            "max_column": column_count,
            "header_row": header_row,
            "header_candidate_row": sheet_summary["header_candidate_row"],
            "header_confirmed": bool(header_row),
            "data_rows_estimated": data_rows,
        },
        "columns": [
            {
                "index": column["index"],
                "letter": column["letter"],
                "technical_id": column["technical_id"],
                "display_name": column["display_name"],
                "confirmed_name": column["confirmed_name"],
                "header_confirmed": column["header_confirmed"],
            }
            for column in columns
        ],
        "quality_by_column": column_stats,
        "critical_completeness_issues": critical_issues,
        "business_patterns": {
            "frequency": build_counter_summary(columns, frequency_by_column, "vigencia/frecuencia"),
            "policy_number": build_counter_summary(columns, policy_by_column, "numero de poliza"),
            "identification_formats": build_counter_summary(
                columns, identification_by_column, "identificacion"
            ),
            "separate_due_date": detect_separate_due_date_columns(columns),
            "detail_field": build_presence_summary(columns, detail_counts, "detalle"),
            "plate_farm_field": build_presence_summary(columns, plate_farm_counts, "placa/finca"),
        },
    }


def build_counter_summary(
    columns: list[dict[str, Any]], counters_by_column: dict[int, Counter], label: str
) -> dict[str, Any]:
    aggregate = Counter()
    output_columns = []

    for index, counter in counters_by_column.items():
        aggregate.update(counter)
        column = columns[index]
        output_columns.append(
            {
                "index": index + 1,
                "technical_id": column["technical_id"],
                "display_name": column["display_name"],
                "header_confirmed": column["header_confirmed"],
                "categories": dict(sorted(counter.items())),
            }
        )

    return {
        "label": label,
        "present": bool(counters_by_column),
        "has_dm": aggregate.get(FREQUENCY_DM, 0) > 0 if label == "vigencia/frecuencia" else None,
        "aggregate_categories": dict(sorted(aggregate.items())),
        "columns": output_columns,
    }


def build_presence_summary(
    columns: list[dict[str, Any]], counts_by_column: dict[int, dict[str, int]], label: str
) -> dict[str, Any]:
    return {
        "label": label,
        "present": bool(counts_by_column),
        "columns": [
            {
                "index": index + 1,
                "technical_id": columns[index]["technical_id"],
                "display_name": columns[index]["display_name"],
                "header_confirmed": columns[index]["header_confirmed"],
                "filled_count": counts["filled_count"],
                "empty_count": counts["empty_count"],
            }
            for index, counts in counts_by_column.items()
        ],
    }


def build_findings(audit: dict[str, Any]) -> tuple[list[str], list[str], list[str]]:
    main = audit["main_sheet"]
    patterns = main["business_patterns"]
    findings: list[str] = []
    warnings: list[str] = []
    recommendations: list[str] = []

    if len(audit["workbook"]["sheets"]) > 1:
        warnings.append(
            "La hoja principal fue seleccionada por volumen de datos y encabezados; conviene validarla manualmente."
        )

    if not main["dimensions"]["header_confirmed"]:
        warnings.append(
            "No se confirmo una fila de encabezados confiable; el reporte usa identificadores tecnicos COL_A, COL_B, etc."
        )

    if patterns["frequency"]["present"]:
        findings.append("Se detectaron columnas candidatas de vigencia o frecuencia.")
    else:
        warnings.append("No se detectaron columnas candidatas de vigencia o frecuencia.")

    if patterns["frequency"]["has_dm"]:
        findings.append("Se detecto presencia de vigencia D.M. o deduccion mensual.")

    if patterns["policy_number"]["present"]:
        findings.append("Se detectaron columnas candidatas de numero de poliza.")
    else:
        warnings.append("No se detectaron columnas candidatas de numero de poliza.")

    if patterns["separate_due_date"]["has_separate_due_date_fields"]:
        findings.append("Se detectaron campos separados para dia, mes y ano de vencimiento.")
    else:
        warnings.append("No se confirmaron campos separados completos para dia, mes y ano de vencimiento.")

    if patterns["detail_field"]["present"]:
        findings.append("Se detecto columna candidata de detalle para anotaciones o relaciones operativas.")

    if patterns["plate_farm_field"]["present"]:
        findings.append("Se detecto columna candidata de placa o finca.")

    if main["critical_completeness_issues"]:
        warnings.append("Hay columnas criticas preliminares con valores vacios; revisar antes de modelar datos.")

    recommendations.extend(
        [
            "Validar con una persona responsable cual hoja debe considerarse principal.",
            "Validar manualmente la fila de encabezados antes de disenar el dataset mejorado.",
            "Disenar el dataset mejorado conservando columnas originales internamente y agregando campos derivados por separado.",
            "Definir un catalogo revisado de vigencias/frecuencias a partir de las categorias observadas.",
            "Confirmar la regla de moneda por prefijos 01/02 y la excepcion de riesgos del trabajo.",
            "Definir reglas de composicion para fechas de vencimiento separadas en dia, mes y ano.",
            "Mantener los reportes de auditoria en data/output/auditoria/ o una ruta local ignorada por Git.",
        ]
    )

    return findings, warnings, recommendations


def build_audit(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet_summaries = [summarize_sheet(sheet) for sheet in workbook.worksheets]
        main_sheet_summary = select_main_sheet(sheet_summaries)
        main_worksheet = workbook[main_sheet_summary["name"]]
        main_sheet = analyze_main_sheet(main_worksheet, main_sheet_summary)

        audit = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "privacy": {
                "safe_report": True,
                "contains_row_samples": False,
                "contains_sensitive_values": False,
                "uses_safe_column_labels_when_uncertain": True,
                "notes": [
                    "No se incluyen muestras de filas.",
                    "No se incluyen nombres, identificaciones, polizas completas ni placas.",
                    "Los encabezados no confirmados o riesgosos se reportan como COL_A, COL_B, etc.",
                    "Los nombres originales de columnas solo se muestran cuando el encabezado fue confirmado y saneado.",
                ],
            },
            "workbook": {
                "file_name": input_path.name,
                "sheets": sheet_summaries,
                "main_sheet_selected": main_sheet_summary["name"],
            },
            "main_sheet": main_sheet,
            "assumptions": [
                "La hoja principal se estima por cantidad de filas de datos y encabezados seguros detectados.",
                "Si no hay encabezado confiable, el analisis usa identificadores tecnicos de columna.",
                "Las categorias de vigencia/frecuencia son observadas y preliminares.",
                "La moneda por prefijos 01/02 es preliminar y no aplica a polizas completamente numericas.",
                "Las clasificaciones de identificacion son probables, no validaciones definitivas.",
            ],
        }
        findings, warnings, recommendations = build_findings(audit)
        audit["integrity_findings"] = findings
        audit["warnings"] = warnings
        audit["recommendations"] = recommendations
        return audit
    finally:
        workbook.close()


def markdown_table(headers: list[str], rows: list[list[Any]]) -> str:
    output = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        output.append("| " + " | ".join(str(value) for value in row) + " |")
    return "\n".join(output)


def format_header_row(value: Any) -> str:
    return str(value) if value else "No confirmado"


def build_markdown_report(audit: dict[str, Any]) -> str:
    workbook = audit["workbook"]
    main = audit["main_sheet"]
    patterns = main["business_patterns"]

    sheet_rows = [
        [
            sheet["name"],
            sheet["max_row"],
            sheet["max_column"],
            format_header_row(sheet["header_row"]),
            "si" if sheet["header_confirmed"] else "no",
            sheet["data_rows_estimated"],
        ]
        for sheet in workbook["sheets"]
    ]
    column_rows = [
        [
            column["index"],
            column["letter"],
            column["technical_id"],
            column["display_name"],
            "si" if column["header_confirmed"] else "no",
            column["filled_count"],
            column["empty_count"],
            f'{column["empty_percentage"]}%',
            "si" if column["critical"] else "no",
        ]
        for column in main["quality_by_column"]
    ]

    report = [
        "# Resumen de Auditoria Local",
        "",
        "Reporte estructural seguro. No incluye muestras de filas ni valores sensibles.",
        "",
        "## Privacidad",
        "",
        "- No se incluyen nombres de clientes.",
        "- No se incluyen identificaciones.",
        "- No se incluyen numeros completos de poliza.",
        "- No se incluyen placas, fincas ni anotaciones reales.",
        "- Los encabezados dudosos o no confirmados se muestran como `COL_A`, `COL_B`, etc.",
        "",
        "## Workbook",
        "",
        f"- Archivo analizado: `{workbook['file_name']}`",
        f"- Hoja principal seleccionada: `{workbook['main_sheet_selected']}`",
        f"- Generado: `{audit['generated_at']}`",
        "",
        "## Hojas Detectadas",
        "",
        markdown_table(
            [
                "Hoja",
                "Filas max.",
                "Columnas max.",
                "Fila encabezado",
                "Encabezado confirmado",
                "Filas de datos estimadas",
            ],
            sheet_rows,
        ),
        "",
        "## Dimensiones de Hoja Principal",
        "",
        f"- Filas maximas: `{main['dimensions']['max_row']}`",
        f"- Columnas maximas: `{main['dimensions']['max_column']}`",
        f"- Fila de encabezados: `{format_header_row(main['dimensions']['header_row'])}`",
        f"- Encabezado confirmado: `{main['dimensions']['header_confirmed']}`",
        f"- Filas de datos estimadas: `{main['dimensions']['data_rows_estimated']}`",
        "",
        "## Calidad por Columna",
        "",
        markdown_table(
            [
                "#",
                "Letra",
                "ID tecnico",
                "Etiqueta segura",
                "Encabezado confirmado",
                "Con dato",
                "Vacias",
                "% vacio",
                "Critica preliminar",
            ],
            column_rows,
        ),
        "",
        "## Vigencias y Frecuencias",
        "",
        format_category_block(patterns["frequency"]),
        "",
        "## Patrones de Numero de Poliza",
        "",
        format_category_block(patterns["policy_number"]),
        "",
        "## Formatos de Identificacion",
        "",
        format_category_block(patterns["identification_formats"]),
        "",
        "## Fecha de Vencimiento Separada",
        "",
        f"- Campos completos dia/mes/ano detectados: `{patterns['separate_due_date']['has_separate_due_date_fields']}`",
        "",
        "## Campo Detalle",
        "",
        format_presence_block(patterns["detail_field"]),
        "",
        "## Campo Placa / Finca",
        "",
        format_presence_block(patterns["plate_farm_field"]),
        "",
        "## Hallazgos de Integridad",
        "",
        format_bullets(audit["integrity_findings"]),
        "",
        "## Advertencias",
        "",
        format_bullets(audit["warnings"]),
        "",
        "## Recomendaciones",
        "",
        format_bullets(audit["recommendations"]),
        "",
        "## Supuestos",
        "",
        format_bullets(audit["assumptions"]),
        "",
    ]
    return "\n".join(report)


def format_category_block(summary: dict[str, Any]) -> str:
    if not summary["present"]:
        return "- No se detectaron columnas candidatas."

    lines = ["Categorias agregadas:"]
    for category, count in summary["aggregate_categories"].items():
        lines.append(f"- `{category}`: `{count}`")
    lines.append("")
    lines.append("Columnas candidatas:")
    for column in summary["columns"]:
        lines.append(
            f"- {column['technical_id']} `{column['display_name']}` "
            f"(encabezado confirmado: {column['header_confirmed']})"
        )
    return "\n".join(lines)


def format_presence_block(summary: dict[str, Any]) -> str:
    if not summary["present"]:
        return "- No se detectaron columnas candidatas."

    lines = []
    for column in summary["columns"]:
        lines.append(
            f"- {column['technical_id']} `{column['display_name']}`: "
            f"{column['filled_count']} con dato, {column['empty_count']} vacias."
        )
    return "\n".join(lines)


def format_bullets(items: list[str]) -> str:
    if not items:
        return "- Sin hallazgos registrados."
    return "\n".join(f"- {item}" for item in items)


def write_audit_outputs(audit: dict[str, Any], output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "resumen_auditoria.json"
    markdown_path = output_dir / "resumen_auditoria.md"

    json_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(build_markdown_report(audit), encoding="utf-8")
    return markdown_path, json_path


def audit_workbook(input_path: str | Path, output_dir: str | Path) -> dict[str, Any]:
    input_path = Path(input_path)
    output_dir = Path(output_dir)

    if not input_path.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("El archivo de entrada debe ser .xlsx o .xlsm.")

    audit = build_audit(input_path)
    markdown_path, json_path = write_audit_outputs(audit, output_dir)
    audit["output_files"] = {
        "markdown": str(markdown_path),
        "json": str(json_path),
    }
    return audit


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera una auditoria estructural segura de un workbook local confidencial."
    )
    parser.add_argument("input_path", help="Ruta del archivo Excel local.")
    parser.add_argument("output_dir", help="Carpeta local donde se guardara la auditoria.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    audit = audit_workbook(args.input_path, args.output_dir)
    print("Auditoria segura generada.")
    print(f"- {audit['output_files']['markdown']}")
    print(f"- {audit['output_files']['json']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
