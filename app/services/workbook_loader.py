"""Lector controlado del archivo operativo Control Cartera.

Este servicio carga datos del Control Cartera en memoria sin guardar el archivo
de origen, crear reportes ni imprimir valores de filas. La fuente activa es el
archivo operativo en `data/input/`, y las columnas auxiliares técnicas quedan
fuera del flujo visible.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import WorkbookLoadError
from app.core.paths import get_project_paths
from app.domain.workbook_records import (
    WorkbookColumn,
    WorkbookLoadResult,
    WorkbookLoadSummary,
    WorkbookRowRecord,
)
from app.domain.workbook_rules import normalize_text


MAIN_SHEET_NAME = "CONTROLCARTERA"
DEFAULT_CONTROL_CARTERA_FILENAME = "CONTROLCARTERA_V2.xlsx"
SYSTEM_AUXILIARY_PREFIXES = ("GS_",)
HEADER_KEYWORDS = {
    "ano",
    "asegurado",
    "cedula",
    "cliente",
    "detalle",
    "dia",
    "finca",
    "identificacion",
    "mes",
    "placa",
    "poliza",
    "venc",
    "vigencia",
}


def get_default_control_cartera_path(project_root: str | Path | None = None) -> Path:
    """Devuelve la ruta local normal del Control Cartera operativo."""
    root = Path(project_root).resolve() if project_root is not None else None
    return get_project_paths(root).data_input_dir / DEFAULT_CONTROL_CARTERA_FILENAME


def load_control_cartera(input_path: str | Path) -> WorkbookLoadResult:
    """Carga un Control Cartera en modo de solo lectura.

    La función no modifica, guarda ni crea archivos. Los valores reales de filas
    se almacenan solo en registros en memoria para la GUI local.
    """
    control_path = _validate_workbook_path(input_path)
    workbook = load_workbook(control_path, read_only=True, data_only=False)
    try:
        if MAIN_SHEET_NAME not in workbook.sheetnames:
            raise WorkbookLoadError(f"No existe la hoja requerida: {MAIN_SHEET_NAME}")

        worksheet = workbook[MAIN_SHEET_NAME]
        header_row = _detect_header_row(worksheet)
        columns = _read_columns(worksheet, header_row)
        if not columns:
            raise WorkbookLoadError("No se detectaron columnas visibles en el Control Cartera.")

        relevant_indexes = _detect_operational_columns(columns)
        records, rows_skipped = _load_records(worksheet, header_row, columns, relevant_indexes)
        summary = _build_summary(control_path, worksheet, header_row, columns, records, rows_skipped)
        return WorkbookLoadResult(summary=summary, records=tuple(records))
    finally:
        workbook.close()


def _validate_workbook_path(input_path: str | Path) -> Path:
    control_path = Path(input_path).resolve()
    if not control_path.exists():
        raise WorkbookLoadError("El archivo seleccionado no existe.")
    if not control_path.is_file():
        raise WorkbookLoadError("La ruta seleccionada no corresponde a un archivo.")
    if control_path.suffix.lower() != ".xlsx":
        raise WorkbookLoadError("Formato no admitido. Seleccione un archivo Excel con extensión .xlsx.")
    return control_path


def _detect_header_row(worksheet: Worksheet, scan_rows: int = 50) -> int:
    """Detecta la fila de encabezados más probable a partir de nombres de columnas originales."""
    best_row = 1
    best_score = -1
    max_scan = min(worksheet.max_row or 1, scan_rows)

    for row_index in range(1, max_scan + 1):
        score = _score_header_row(worksheet, row_index)
        if score > best_score:
            best_score = score
            best_row = row_index

    return best_row


def _score_header_row(worksheet: Worksheet, row_index: int) -> int:
    score = 0
    for cell in worksheet[row_index]:
        header_text = str(cell.value).strip() if cell.value is not None else ""
        if _is_system_auxiliary_column(header_text):
            continue
        normalized = normalize_text(header_text)
        if not normalized:
            continue
        score += 1
        if any(keyword in normalized for keyword in HEADER_KEYWORDS):
            score += 8
    return score


def _read_columns(worksheet: Worksheet, header_row: int) -> list[WorkbookColumn]:
    header_values = [cell.value for cell in worksheet[header_row]]
    columns: list[WorkbookColumn] = []
    used_display_names: set[str] = set()
    last_header_column = _last_non_empty_header_column(header_values)

    for index in range(1, last_header_column + 1):
        raw_header = header_values[index - 1] if index <= len(header_values) else None
        header_text = str(raw_header).strip() if raw_header is not None else ""
        if _is_system_auxiliary_column(header_text):
            continue
        if not header_text and not _column_has_sampled_data(worksheet, header_row, index):
            continue

        technical_name = f"COL_{get_column_letter(index)}"
        fallback_name = f"Columna {get_column_letter(index)}"
        display_name = _unique_display_name(header_text or fallback_name, used_display_names)
        columns.append(
            WorkbookColumn(
                index=index,
                technical_name=technical_name,
                display_name=display_name,
                original_header=header_text or None,
            )
        )

    return columns


def _detect_operational_columns(columns: list[WorkbookColumn]) -> set[int]:
    relevant = {
        column.index
        for column in columns
        if any(keyword in normalize_text(column.original_header) for keyword in HEADER_KEYWORDS)
    }
    if relevant:
        return relevant
    return {column.index for column in columns}


def _load_records(
    worksheet: Worksheet,
    header_row: int,
    columns: list[WorkbookColumn],
    relevant_indexes: set[int],
) -> tuple[list[WorkbookRowRecord], int]:
    records: list[WorkbookRowRecord] = []
    rows_skipped = 0
    column_indexes = [column.index for column in columns]

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=worksheet.max_row,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        if not _row_has_operational_data(row, column_indexes, relevant_indexes):
            rows_skipped += 1
            continue

        values_by_column: dict[str, Any] = {}
        for column in columns:
            value = row[column.index - 1] if column.index <= len(row) else None
            values_by_column[column.display_name] = value

        records.append(
            WorkbookRowRecord(
                row_number=row_number,
                values_by_column=values_by_column,
            )
        )

    return records, rows_skipped


def _build_summary(
    control_path: Path,
    worksheet: Worksheet,
    header_row: int,
    columns: list[WorkbookColumn],
    records: list[WorkbookRowRecord],
    rows_skipped: int,
) -> WorkbookLoadSummary:
    return WorkbookLoadSummary(
        source_name=control_path.name,
        sheet_name=worksheet.title,
        header_row=header_row,
        total_rows=worksheet.max_row or 0,
        total_columns=len(columns),
        useful_rows_detected=len(records),
        records_loaded=len(records),
        rows_skipped=rows_skipped,
        detected_columns=tuple(column.display_name for column in columns),
        visible_columns=tuple(column.display_name for column in columns),
        read_only=True,
        warnings=_build_warnings(),
    )


def _build_warnings() -> tuple[str, ...]:
    return (
        "Carga de solo lectura; no se modificó ni guardó el Control Cartera.",
        "Los registros quedan solo en memoria para visualización local.",
    )


def _row_has_operational_data(
    row: tuple[Any, ...],
    visible_indexes: list[int],
    relevant_indexes: set[int],
) -> bool:
    return any(
        _has_value(row[index - 1] if index <= len(row) else None)
        for index in visible_indexes
        if index in relevant_indexes
    )


def _has_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or value.strip() != "")


def _is_system_auxiliary_column(header_text: str) -> bool:
    return any(header_text.upper().startswith(prefix) for prefix in SYSTEM_AUXILIARY_PREFIXES)


def _last_non_empty_header_column(header_values: list[Any]) -> int:
    last_index = 0
    for index, value in enumerate(header_values, start=1):
        header_text = str(value).strip() if value is not None else ""
        if header_text and not _is_system_auxiliary_column(header_text):
            last_index = index
    return last_index


def _column_has_sampled_data(
    worksheet: Worksheet,
    header_row: int,
    column_index: int,
    sample_rows: int = 50,
) -> bool:
    max_row = min(worksheet.max_row or header_row, header_row + sample_rows)
    for row_index in range(header_row + 1, max_row + 1):
        if _has_value(worksheet.cell(row=row_index, column=column_index).value):
            return True
    return False


def _unique_display_name(display_name: str, used_display_names: set[str]) -> str:
    if display_name not in used_display_names:
        used_display_names.add(display_name)
        return display_name

    counter = 2
    candidate = f"{display_name}_{counter}"
    while candidate in used_display_names:
        counter += 1
        candidate = f"{display_name}_{counter}"
    used_display_names.add(candidate)
    return candidate
