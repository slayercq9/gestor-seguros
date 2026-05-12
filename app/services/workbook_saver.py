"""Guardado seguro del Control Cartera.

El servicio escribe cambios aplicados en memoria usando fila real e índice real
de columna. `Guardar como` exporta una copia y `Guardar` modifica el archivo
cargado solo después de crear un respaldo automático.
"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.core.exceptions import WorkbookSaveError
from app.core.paths import get_project_paths
from app.domain.column_standards import ISSUE_DATE, resolve_column_key
from app.services.workbook_loader import MAIN_SHEET_NAME, _detect_header_row, _read_columns

DIRECT_SAVE_BACKUP_DIR_NAME = "guardado_control_cartera"


@dataclass(frozen=True)
class WorkbookCellUpdate:
    """Cambio puntual que debe escribirse en una copia del Control Cartera."""

    row_number: int
    column_name: str
    value: object
    column_index: int | None = None


def save_control_cartera_as(
    source_path: str | Path,
    destination_path: str | Path,
    updates: Iterable[WorkbookCellUpdate],
    *,
    sheet_name: str = MAIN_SHEET_NAME,
    header_row: int | None = None,
) -> Path:
    """Guarda una copia `.xlsx` aplicando cambios de memoria.

    El archivo fuente se abre solo como plantilla de lectura/escritura en memoria
    y se guarda exclusivamente en `destination_path`.
    """
    source = _validate_source_path(source_path)
    destination = _validate_destination_path(source, destination_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookSaveError(f"No existe la hoja requerida: {sheet_name}")

        worksheet = workbook[sheet_name]
        effective_header_row = header_row or _detect_header_row(worksheet)
        column_map = _column_index_by_display_name(worksheet, effective_header_row)
        for update in updates:
            column_index = _resolve_update_column_index(update, column_map)
            worksheet.cell(
                row=update.row_number,
                column=column_index,
                value=_value_for_save(update.column_name, update.value),
            )

        workbook.save(destination)
    finally:
        workbook.close()

    return destination


def save_control_cartera(
    source_path: str | Path,
    updates: Iterable[WorkbookCellUpdate],
    *,
    sheet_name: str = MAIN_SHEET_NAME,
    header_row: int | None = None,
    backup_dir: str | Path | None = None,
) -> Path:
    """Guarda cambios sobre el archivo cargado después de crear respaldo.

    Devuelve la ruta del respaldo creado. Si el respaldo falla, no se intenta
    escribir sobre el archivo cargado.
    """
    source = _validate_source_path(source_path)
    backup_path = _create_backup(source, backup_dir)
    _write_updates(source, source, updates, sheet_name=sheet_name, header_row=header_row)
    return backup_path


def _validate_source_path(source_path: str | Path) -> Path:
    source = Path(source_path).resolve()
    if not source.exists() or not source.is_file():
        raise WorkbookSaveError("El archivo fuente no existe.")
    if source.suffix.lower() != ".xlsx":
        raise WorkbookSaveError("El archivo fuente debe tener extension .xlsx.")
    return source


def _validate_destination_path(source: Path, destination_path: str | Path) -> Path:
    destination = Path(destination_path).resolve()
    if destination.suffix.lower() != ".xlsx":
        raise WorkbookSaveError("La ruta de destino debe tener extension .xlsx.")
    if destination == source:
        raise WorkbookSaveError("Guardar como no puede sobrescribir el archivo cargado.")
    return destination


def _write_updates(
    source: Path,
    destination: Path,
    updates: Iterable[WorkbookCellUpdate],
    *,
    sheet_name: str,
    header_row: int | None,
) -> None:
    workbook = load_workbook(source)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookSaveError(f"No existe la hoja requerida: {sheet_name}")

        worksheet = workbook[sheet_name]
        effective_header_row = header_row or _detect_header_row(worksheet)
        column_map = _column_index_by_display_name(worksheet, effective_header_row)
        for update in updates:
            column_index = _resolve_update_column_index(update, column_map)
            worksheet.cell(
                row=update.row_number,
                column=column_index,
                value=_value_for_save(update.column_name, update.value),
            )

        workbook.save(destination)
    finally:
        workbook.close()


def _create_backup(source: Path, backup_dir: str | Path | None) -> Path:
    target_dir = Path(backup_dir).resolve() if backup_dir is not None else _default_backup_dir()
    target_dir.mkdir(parents=True, exist_ok=True)
    backup_path = _unique_backup_path(target_dir)
    try:
        shutil.copy2(source, backup_path)
    except OSError as exc:
        raise WorkbookSaveError("No fue posible crear el respaldo automático antes de guardar.") from exc
    return backup_path


def _default_backup_dir() -> Path:
    return get_project_paths().data_backups_dir / DIRECT_SAVE_BACKUP_DIR_NAME


def _unique_backup_path(backup_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"CONTROLCARTERA_V2_backup_{timestamp}"
    candidate = backup_dir / f"{base_name}.xlsx"
    counter = 2
    while candidate.exists():
        candidate = backup_dir / f"{base_name}_{counter}.xlsx"
        counter += 1
    return candidate


def _column_index_by_display_name(worksheet: object, header_row: int) -> dict[str, int]:
    columns = _read_columns(worksheet, header_row)
    return {column.display_name: column.index for column in columns}


def _resolve_update_column_index(update: WorkbookCellUpdate, column_map: dict[str, int]) -> int:
    """Resuelve la columna real del Excel sin depender de valores editables."""
    if update.row_number < 1:
        raise WorkbookSaveError("Falta metadata de fila para guardar el cambio.")
    if update.column_index is not None:
        if update.column_index < 1 or update.column_index not in set(column_map.values()):
            raise WorkbookSaveError("Falta metadata de columna para guardar el cambio.")
        return update.column_index
    if update.column_name not in column_map:
        raise WorkbookSaveError(f"Falta metadata de columna para guardar: {update.column_name}")
    return column_map[update.column_name]


def _value_for_save(column_name: str, value: object) -> object:
    if resolve_column_key(column_name) != ISSUE_DATE:
        return value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    match = re.fullmatch(r"(\d{4}-\d{2}-\d{2})(?:[ T].*)?", text)
    if match:
        return match.group(1)
    return text
