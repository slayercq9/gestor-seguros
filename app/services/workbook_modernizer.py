"""Controlled workbook modernization service.

The service creates a local modernized copy and structural reports. It never
saves over the original workbook, never deletes records, and does not add
auxiliary data columns to the operational sheet.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import WorkbookModernizationError
from app.domain.workbook_rules import normalize_text, safe_text


REPORT_MD_NAME = "resumen_modernizacion.md"
REPORT_JSON_NAME = "resumen_modernizacion.json"
CONTROL_SHEET_NAME = "CONTROL_MODERNIZACION"

HEADER_KEYWORDS = {
    "ano",
    "asegurado",
    "cedula",
    "cliente",
    "detalle",
    "dia",
    "finca",
    "identificacion",
    "mes",
    "placa",
    "poliza",
    "venc",
    "vigencia",
}


@dataclass(frozen=True)
class ModernizationResult:
    """Paths and metadata produced by workbook modernization."""

    output_workbook: Path
    markdown_report: Path
    json_report: Path
    main_sheet: str
    useful_rows: int
    rows_skipped: int
    visible_columns: tuple[str, ...]


def modernize_workbook(input_path: str | Path, output_dir: str | Path) -> ModernizationResult:
    """Create a modernized local copy and local control reports.

    The operational sheet keeps its original columns and values. Modernization
    is limited to non-destructive visual formatting plus local control reports.
    """
    source = Path(input_path).resolve()
    destination_dir = Path(output_dir).resolve()
    if destination_dir == source:
        raise WorkbookModernizationError("La carpeta de salida no puede ser el archivo original.")

    generated_at = datetime.now()
    output_workbook = destination_dir / _build_output_workbook_name(source, generated_at)

    _validate_paths(source, output_workbook)
    destination_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source)
    try:
        main_sheet = _select_main_sheet(workbook.worksheets)
        worksheet = workbook[main_sheet]
        header_row = _detect_header_row(worksheet)
        removed_legacy_columns = _remove_legacy_gs_columns(worksheet, header_row)
        visible_columns = _read_visible_headers(worksheet, header_row)
        useful_rows, rows_skipped = _count_useful_rows(worksheet, header_row, visible_columns)

        _apply_visual_formatting(worksheet, header_row, len(visible_columns))
        _write_control_sheet(
            workbook,
            source,
            output_workbook,
            main_sheet,
            header_row,
            useful_rows,
            rows_skipped,
            visible_columns,
            removed_legacy_columns,
        )

        workbook.save(output_workbook)
    finally:
        workbook.close()

    report = _build_report(
        source,
        output_workbook,
        main_sheet,
        worksheet.max_row,
        len(visible_columns),
        header_row,
        useful_rows,
        rows_skipped,
        visible_columns,
        removed_legacy_columns,
        generated_at,
    )
    markdown_report = destination_dir / REPORT_MD_NAME
    json_report = destination_dir / REPORT_JSON_NAME
    markdown_report.write_text(_render_markdown_report(report), encoding="utf-8")
    json_report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    return ModernizationResult(
        output_workbook=output_workbook,
        markdown_report=markdown_report,
        json_report=json_report,
        main_sheet=main_sheet,
        useful_rows=useful_rows,
        rows_skipped=rows_skipped,
        visible_columns=tuple(visible_columns),
    )


def _build_output_workbook_name(source: Path, generated_at: datetime) -> str:
    timestamp = generated_at.strftime("%Y%m%d_%H%M%S")
    return f"{source.stem}_modernizado_{timestamp}.xlsx"


def _validate_paths(source: Path, output_workbook: Path) -> None:
    if not source.exists():
        raise WorkbookModernizationError(f"No existe el workbook de entrada: {source}")
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise WorkbookModernizationError("El workbook de entrada debe ser .xlsx o .xlsm.")
    if source == output_workbook.resolve():
        raise WorkbookModernizationError("La salida no puede ser el mismo archivo original.")


def _select_main_sheet(worksheets: list[Worksheet]) -> str:
    """Select the sheet with the largest useful data footprint."""
    return max(
        worksheets,
        key=lambda sheet: (
            _count_non_empty_rows(sheet),
            sheet.max_column or 0,
            _score_header_row(sheet, _detect_header_row(sheet)),
        ),
    ).title


def _count_non_empty_rows(worksheet: Worksheet) -> int:
    return sum(1 for row in worksheet.iter_rows(values_only=True) if any(not _is_empty(value) for value in row))


def _detect_header_row(worksheet: Worksheet, scan_rows: int = 20) -> int:
    best_row = 1
    best_score = -1
    max_scan = min(worksheet.max_row or 1, scan_rows)
    for row_index in range(1, max_scan + 1):
        score = _score_header_row(worksheet, row_index)
        if score > best_score:
            best_score = score
            best_row = row_index
    return best_row


def _score_header_row(worksheet: Worksheet, row_index: int) -> int:
    score = 0
    for cell in worksheet[row_index]:
        normalized = normalize_text(cell.value)
        if not normalized:
            continue
        if normalized.startswith("gs "):
            continue
        score += 1
        if any(keyword in normalized for keyword in HEADER_KEYWORDS):
            score += 8
    return score


def _read_visible_headers(worksheet: Worksheet, header_row: int) -> list[str]:
    headers: list[str] = []
    used_headers: set[str] = set()
    for cell in worksheet[header_row]:
        header_text = safe_text(cell.value).strip()
        if not header_text or header_text.upper().startswith("GS_"):
            continue
        headers.append(_unique_header(header_text, used_headers))
    return headers


def _remove_legacy_gs_columns(worksheet: Worksheet, header_row: int) -> int:
    """Remove legacy auxiliary columns from the generated copy only."""
    removed = 0
    for column_index in range(worksheet.max_column or 0, 0, -1):
        header_text = safe_text(worksheet.cell(row=header_row, column=column_index).value).strip()
        if header_text.upper().startswith("GS_"):
            worksheet.delete_cols(column_index)
            removed += 1
    return removed


def _count_useful_rows(worksheet: Worksheet, header_row: int, visible_columns: list[str]) -> tuple[int, int]:
    useful_rows = 0
    rows_skipped = 0
    max_column = len(visible_columns)

    for row in worksheet.iter_rows(
        min_row=header_row + 1,
        max_row=worksheet.max_row,
        max_col=max_column,
        values_only=True,
    ):
        if any(not _is_empty(value) for value in row):
            useful_rows += 1
        else:
            rows_skipped += 1
    return useful_rows, rows_skipped


def _apply_visual_formatting(worksheet: Worksheet, header_row: int, visible_column_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    max_column = max(visible_column_count, 1)

    for column_index in range(1, max_column + 1):
        cell = worksheet.cell(row=header_row, column=column_index)
        cell.font = header_font
        cell.fill = header_fill
        worksheet.column_dimensions[get_column_letter(column_index)].width = _column_width(worksheet, column_index)

    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1).coordinate
    last_letter = get_column_letter(max_column)
    worksheet.auto_filter.ref = f"A{header_row}:{last_letter}{worksheet.max_row or header_row}"


def _column_width(worksheet: Worksheet, column_index: int, sample_rows: int = 80) -> int:
    max_length = 12
    max_row = min(worksheet.max_row or 1, sample_rows)
    for row_index in range(1, max_row + 1):
        max_length = max(max_length, len(safe_text(worksheet.cell(row=row_index, column=column_index).value)))
    return min(35, max_length + 4)


def _write_control_sheet(
    workbook: Any,
    source: Path,
    output_workbook: Path,
    main_sheet: str,
    header_row: int,
    useful_rows: int,
    rows_skipped: int,
    visible_columns: list[str],
    removed_legacy_columns: int,
) -> None:
    if CONTROL_SHEET_NAME in workbook.sheetnames:
        del workbook[CONTROL_SHEET_NAME]
    sheet = workbook.create_sheet(CONTROL_SHEET_NAME)
    rows = [
        ("concepto", "valor"),
        ("generado_en", datetime.now().isoformat(timespec="seconds")),
        ("archivo_entrada", source.name),
        ("archivo_salida", output_workbook.name),
        ("hoja_principal", main_sheet),
        ("fila_encabezados", header_row),
        ("filas_utiles_detectadas", useful_rows),
        ("filas_omitidas_o_vacias", rows_skipped),
        ("columnas_visibles", len(visible_columns)),
        ("columnas_auxiliares_agregadas", 0),
        ("columnas_auxiliares_heredadas_eliminadas", removed_legacy_columns),
        ("nota", "Copia local modernizada; no reemplaza el workbook original."),
    ]
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 80
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")


def _build_report(
    source: Path,
    output_workbook: Path,
    main_sheet: str,
    rows: int,
    columns: int,
    header_row: int,
    useful_rows: int,
    rows_skipped: int,
    visible_columns: list[str],
    removed_legacy_columns: int,
    generated_at: datetime,
) -> dict[str, Any]:
    warnings = [
        "La hoja principal fue seleccionada automaticamente y debe validarse manualmente.",
        "La modernizacion no agrega columnas auxiliares ni modifica valores originales.",
    ]
    if removed_legacy_columns:
        warnings.append("Se eliminaron columnas auxiliares heredadas de la copia modernizada, no del archivo original.")

    return {
        "generated_at": generated_at.isoformat(timespec="seconds"),
        "input_file": source.name,
        "output_file": str(output_workbook),
        "main_sheet": main_sheet,
        "header_row": header_row,
        "dimensions": {"rows": rows, "visible_columns": columns},
        "columns_added": [],
        "legacy_auxiliary_columns_removed": removed_legacy_columns,
        "useful_rows": useful_rows,
        "rows_skipped": rows_skipped,
        "visible_columns": visible_columns,
        "warnings": warnings,
        "recommendations": [
            "Validar manualmente que la hoja seleccionada sea la cartera operativa.",
            "Mantener el workbook original como respaldo operativo hasta aprobacion humana.",
            "Gestionar reglas de validacion internamente en fases futuras, sin ensuciar la visualizacion.",
        ],
        "privacy": {
            "contains_row_samples": False,
            "contains_sensitive_values": False,
            "notes": [
                "El reporte contiene estadisticas y rutas locales, no valores reales de filas.",
                "La copia modernizada local puede contener datos reales y esta bajo data/output/.",
            ],
        },
    }


def _render_markdown_report(report: dict[str, Any]) -> str:
    return "\n".join(
        [
            "# Reporte de Modernizacion Workbook",
            "",
            "Reporte local de control. No incluye muestras de filas ni valores sensibles.",
            "",
            f"- Generado: `{report['generated_at']}`",
            f"- Archivo de entrada: `{report['input_file']}`",
            f"- Archivo de salida: `{report['output_file']}`",
            f"- Hoja principal: `{report['main_sheet']}`",
            f"- Fila de encabezados: `{report['header_row']}`",
            f"- Filas maximas: `{report['dimensions']['rows']}`",
            f"- Filas utiles detectadas: `{report['useful_rows']}`",
            f"- Filas omitidas o vacias: `{report['rows_skipped']}`",
            f"- Columnas visibles: `{report['dimensions']['visible_columns']}`",
            f"- Columnas agregadas: `{len(report['columns_added'])}`",
            f"- Columnas auxiliares heredadas eliminadas: `{report['legacy_auxiliary_columns_removed']}`",
            "",
            "## Advertencias",
            "",
            _format_bullets(report["warnings"]),
            "",
            "## Recomendaciones",
            "",
            _format_bullets(report["recommendations"]),
            "",
        ]
    )


def _format_bullets(items: list[str]) -> str:
    if not items:
        return "- Sin registros."
    return "\n".join(f"- {item}" for item in items)


def _unique_header(header_text: str, used_headers: set[str]) -> str:
    if header_text not in used_headers:
        used_headers.add(header_text)
        return header_text

    counter = 2
    candidate = f"{header_text}_{counter}"
    while candidate in used_headers:
        counter += 1
        candidate = f"{header_text}_{counter}"
    used_headers.add(candidate)
    return candidate


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")
