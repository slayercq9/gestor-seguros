import shutil
import subprocess
import sys
import uuid
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.core.exceptions import WorkbookLoadError
from app.services.workbook_loader import (
    DEFAULT_CONTROL_CARTERA_FILENAME,
    MAIN_SHEET_NAME,
    get_default_control_cartera_path,
    load_control_cartera,
)


@contextmanager
def workspace_tempdir():
    base_dir = Path("data/output")
    base_dir.mkdir(parents=True, exist_ok=True)
    path = base_dir / f"tmp-loader-{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def build_control_cartera(
    path: Path,
    include_sheet: bool = True,
    include_system_columns: bool = False,
    include_coverage_columns: bool = False,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = MAIN_SHEET_NAME if include_sheet else "OTRA_HOJA"

    headers = ["Cliente", "Poliza", "Vigencia", "Dia", "Mes", "Ano"]
    if include_coverage_columns:
        headers.extend(["Cobertura A", "Cobertura Especial"])
    if include_system_columns:
        headers.extend(["GS_ES_DM", "GS_REQUIERE_REVISION"])
    worksheet.append(headers)
    row_a = ["Registro Ficticio A", "POL-FICT-001", "D.M.", 1, 5, 2026]
    row_b = ["Registro Ficticio B", "POL-FICT-002", "Anual", 15, 8, 2026]
    if include_coverage_columns:
        row_a.extend(["Cobertura Ficticia A", "Cobertura Ficticia B"])
        row_b.extend(["Cobertura Ficticia C", None])
    if include_system_columns:
        row_a.extend(["si", "no"])
        row_b.extend(["no", "no"])
    worksheet.append(row_a)
    worksheet.append(row_b)
    worksheet.cell(row=20, column=1).fill = PatternFill("solid", fgColor="FFFFFF")
    workbook.save(path)


def test_carga_control_cartera_ficticio_sin_modificar_fuente():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        build_control_cartera(source)
        original_bytes = source.read_bytes()

        result = load_control_cartera(source)

        assert source.read_bytes() == original_bytes
        assert result.summary.sheet_name == MAIN_SHEET_NAME
        assert result.summary.read_only is True
        assert result.summary.useful_rows_detected == 2
        assert result.summary.records_loaded == 2
        assert result.summary.rows_skipped >= 1
        assert result.summary.visible_columns == ("Cliente", "Poliza", "Vigencia", "Dia", "Mes", "Ano")
        assert not any(column.startswith("GS_") for column in result.summary.visible_columns)
        assert result.records[0].values_by_column["Cliente"] == "Registro Ficticio A"


def test_resuelve_ruta_predeterminada_de_input():
    default_path = get_default_control_cartera_path()

    assert default_path.name == DEFAULT_CONTROL_CARTERA_FILENAME
    assert default_path.parent.name == "input"
    assert default_path.parent.parent.name == "data"


def test_valida_archivo_inexistente_y_extension_incorrecta():
    with workspace_tempdir() as temp_dir:
        with pytest.raises(WorkbookLoadError):
            load_control_cartera(temp_dir / "no_existe.xlsx")

        invalid = temp_dir / "control_cartera.txt"
        invalid.write_text("no es excel", encoding="utf-8")

        with pytest.raises(WorkbookLoadError):
            load_control_cartera(invalid)


def test_requiere_hoja_principal_controles():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "sin_hoja.xlsx"
        build_control_cartera(source, include_sheet=False)

        with pytest.raises(WorkbookLoadError):
            load_control_cartera(source)


def test_ignora_columnas_auxiliares_tecnicas_si_existen():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_con_columnas_tecnicas.xlsx"
        build_control_cartera(source, include_system_columns=True)

        result = load_control_cartera(source)

        assert result.summary.records_loaded == 2
        assert "GS_ES_DM" not in result.summary.visible_columns
        assert "GS_REQUIERE_REVISION" not in result.summary.visible_columns
        assert "GS_ES_DM" not in result.records[0].values_by_column


def test_oculta_coberturas_pero_las_conserva_en_memoria():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_con_coberturas.xlsx"
        build_control_cartera(source, include_coverage_columns=True)

        result = load_control_cartera(source)

        assert "Cobertura A" in result.summary.detected_columns
        assert "Cobertura Especial" in result.summary.detected_columns
        assert "Cobertura A" not in result.summary.visible_columns
        assert "Cobertura Especial" not in result.summary.visible_columns
        assert result.records[0].values_by_column["Cobertura A"] == "Cobertura Ficticia A"
        assert result.records[0].values_by_column["Cobertura Especial"] == "Cobertura Ficticia B"


def test_script_imprime_resumen_sin_valores_sensibles_ficticios():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        build_control_cartera(source)

        completed = subprocess.run(
            [sys.executable, "scripts/cargar_control_cartera.py", str(source)],
            check=False,
            capture_output=True,
            text=True,
        )

        assert completed.returncode == 0
        assert "Carga controlada completada" in completed.stdout
        assert "Filas utiles detectadas" in completed.stdout
        assert "Columnas visibles" in completed.stdout
        assert "Columnas GS" not in completed.stdout
        assert "Registro Ficticio A" not in completed.stdout
        assert "POL-FICT-001" not in completed.stdout


def test_no_crea_outputs_permanentes():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        build_control_cartera(source)
        before = {path.name for path in temp_dir.iterdir()}

        load_control_cartera(source)

        after = {path.name for path in temp_dir.iterdir()}
        assert after == before

        workbook = load_workbook(source, read_only=True)
        try:
            assert workbook.sheetnames == [MAIN_SHEET_NAME]
        finally:
            workbook.close()
