import json
import shutil
import uuid
from contextlib import contextmanager
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.limpiar_workbook_operativo import (
    OBSOLETE_SHEET_NAME,
    clean_operational_workbook,
)


@contextmanager
def workspace_tempdir():
    base_dir = Path("data/output")
    base_dir.mkdir(parents=True, exist_ok=True)
    path = base_dir / f"tmp-maintenance-{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def build_workbook(path: Path, include_obsolete_sheet: bool = True) -> None:
    workbook = Workbook()
    main = workbook.active
    main.title = "Cartera"
    main.append(["Cliente", "Poliza"])
    main.append(["Ana Segura", "01-ABC-123"])

    if include_obsolete_sheet:
        obsolete = workbook.create_sheet(OBSOLETE_SHEET_NAME)
        obsolete.append(["contenido obsoleto"])

    workbook.create_sheet("Catalogos")
    workbook.save(path)


def read_main_rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path)
    try:
        return [tuple(row) for row in workbook["Cartera"].iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_crea_respaldo_y_elimina_solo_hoja_obsoleta():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "workbook.xlsx"
        backup_dir = temp_dir / "backups"
        report_dir = temp_dir / "reportes"
        build_workbook(source, include_obsolete_sheet=True)
        original_main_rows = read_main_rows(source)

        result = clean_operational_workbook(source, backup_dir, report_dir)

        assert result.backup_path.exists()
        assert result.sheet_found is True
        assert result.sheet_deleted is True
        assert result.markdown_report.exists()
        assert result.json_report.exists()

        workbook = load_workbook(source)
        try:
            assert OBSOLETE_SHEET_NAME not in workbook.sheetnames
            assert "Cartera" in workbook.sheetnames
            assert "Catalogos" in workbook.sheetnames
        finally:
            workbook.close()

        assert read_main_rows(source) == original_main_rows


def test_si_hoja_no_existe_no_modifica_workbook():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "workbook.xlsx"
        backup_dir = temp_dir / "backups"
        report_dir = temp_dir / "reportes"
        build_workbook(source, include_obsolete_sheet=False)
        original_bytes = source.read_bytes()

        result = clean_operational_workbook(source, backup_dir, report_dir)

        assert result.backup_path.exists()
        assert result.sheet_found is False
        assert result.sheet_deleted is False
        assert source.read_bytes() == original_bytes

        data = json.loads(result.json_report.read_text(encoding="utf-8"))
        assert data["sheet_found"] is False
        assert data["sheet_deleted"] is False
        assert data["privacy"]["contains_row_data"] is False


def test_reporte_no_incluye_datos_de_filas():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "workbook.xlsx"
        backup_dir = temp_dir / "backups"
        report_dir = temp_dir / "reportes"
        build_workbook(source, include_obsolete_sheet=True)

        result = clean_operational_workbook(source, backup_dir, report_dir)
        combined = result.markdown_report.read_text(encoding="utf-8")
        combined += result.json_report.read_text(encoding="utf-8")

        assert "Ana Segura" not in combined
        assert "01-ABC-123" not in combined
        assert OBSOLETE_SHEET_NAME in combined
