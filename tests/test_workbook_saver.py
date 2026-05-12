from datetime import datetime
import shutil
import uuid
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.core.exceptions import WorkbookSaveError
from app.services.workbook_loader import MAIN_SHEET_NAME, load_control_cartera
from app.services.workbook_saver import WorkbookCellUpdate, save_control_cartera, save_control_cartera_as


@contextmanager
def workspace_tempdir():
    base_dir = Path("data/output")
    base_dir.mkdir(parents=True, exist_ok=True)
    path = base_dir / f"tmp-saver-{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def build_control_cartera(path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = MAIN_SHEET_NAME
    worksheet.append(["Nº Póliza", "Nombre del Asegurado", "Emisión", "Vigencia", "Cobertura A", "Detalle"])
    worksheet.append(["01-FICT-001", "Persona Ficticia A", datetime(2022, 3, 8, 0, 0, 0), "Anual", "Cobertura Ficticia", "Detalle A"])
    worksheet.append(["01-FICT-002", "Persona Ficticia B", datetime(2023, 4, 9, 0, 0, 0), "D.M.", "Cobertura Ficticia B", "Detalle B"])
    extra = workbook.create_sheet("Hoja auxiliar")
    extra["A1"] = "Dato auxiliar"
    workbook.save(path)


def test_guardar_como_crea_copia_sin_modificar_fuente():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        destination = temp_dir / "control_cartera_editado.xlsx"
        build_control_cartera(source)
        original_bytes = source.read_bytes()
        result = load_control_cartera(source)

        save_control_cartera_as(
            source,
            destination,
            (
                WorkbookCellUpdate(2, "Nombre del Asegurado", "Persona Ficticia Editada"),
                WorkbookCellUpdate(2, "Emisión", "2024-05-10 00:00:00"),
            ),
            sheet_name=result.summary.sheet_name,
            header_row=result.summary.header_row,
        )

        assert destination.exists()
        assert source.read_bytes() == original_bytes

        saved = load_workbook(destination)
        try:
            worksheet = saved[MAIN_SHEET_NAME]
            assert worksheet["B2"].value == "Persona Ficticia Editada"
            assert worksheet["C2"].value == "2024-05-10"
            assert worksheet["E2"].value == "Cobertura Ficticia"
            assert worksheet["A3"].value == "01-FICT-002"
            assert worksheet["B3"].value == "Persona Ficticia B"
            assert "Hoja auxiliar" in saved.sheetnames
            assert saved["Hoja auxiliar"]["A1"].value == "Dato auxiliar"
        finally:
            saved.close()


def test_guardar_como_permite_cambiar_poliza_y_nombre_por_coordenada_real():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        destination = temp_dir / "control_cartera_editado.xlsx"
        build_control_cartera(source)
        original_bytes = source.read_bytes()

        save_control_cartera_as(
            source,
            destination,
            (
                WorkbookCellUpdate(2, "Nº Póliza", "02-FICT-999", column_index=1),
                WorkbookCellUpdate(2, "Nombre del Asegurado", "Persona Ficticia Renombrada", column_index=2),
            ),
        )

        assert source.read_bytes() == original_bytes
        saved = load_workbook(destination)
        try:
            worksheet = saved[MAIN_SHEET_NAME]
            assert worksheet["A2"].value == "02-FICT-999"
            assert worksheet["B2"].value == "Persona Ficticia Renombrada"
            assert worksheet["E2"].value == "Cobertura Ficticia"
            assert worksheet["A3"].value == "01-FICT-002"
            assert worksheet["B3"].value == "Persona Ficticia B"
        finally:
            saved.close()


def test_guardar_como_no_usa_poliza_ni_nombre_como_identificador_de_fila():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        destination = temp_dir / "control_cartera_editado.xlsx"
        build_control_cartera(source)

        save_control_cartera_as(
            source,
            destination,
            (WorkbookCellUpdate(3, "Nº Póliza", "01-FICT-001", column_index=1),),
        )

        saved = load_workbook(destination)
        try:
            worksheet = saved[MAIN_SHEET_NAME]
            assert worksheet["A2"].value == "01-FICT-001"
            assert worksheet["A3"].value == "01-FICT-001"
            assert worksheet["B3"].value == "Persona Ficticia B"
        finally:
            saved.close()


def test_guardar_como_bloquea_metadata_de_columna_faltante():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        destination = temp_dir / "control_cartera_editado.xlsx"
        build_control_cartera(source)

        with pytest.raises(WorkbookSaveError, match="metadata de columna"):
            save_control_cartera_as(source, destination, (WorkbookCellUpdate(2, "Columna inexistente", "Valor"),))


def test_guardar_como_conserva_columnas_y_filas_no_modificadas():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        destination = temp_dir / "control_cartera_copia.xlsx"
        build_control_cartera(source)

        save_control_cartera_as(source, destination, ())

        source_workbook = load_workbook(source)
        saved_workbook = load_workbook(destination)
        try:
            source_sheet = source_workbook[MAIN_SHEET_NAME]
            saved_sheet = saved_workbook[MAIN_SHEET_NAME]
            assert [cell.value for cell in saved_sheet[1]] == [cell.value for cell in source_sheet[1]]
            assert saved_sheet["E2"].value == source_sheet["E2"].value
            assert saved_sheet["F3"].value == source_sheet["F3"].value
        finally:
            source_workbook.close()
            saved_workbook.close()


def test_guardar_como_bloquea_ruta_igual_a_fuente():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        build_control_cartera(source)

        with pytest.raises(WorkbookSaveError):
            save_control_cartera_as(source, source, ())


def test_guardar_sobre_archivo_cargado_crea_respaldo_y_escribe_cambios():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        backup_dir = temp_dir / "backups"
        build_control_cartera(source)

        backup_path = save_control_cartera(
            source,
            (
                WorkbookCellUpdate(2, "Nº Póliza", "02-FICT-999", column_index=1),
                WorkbookCellUpdate(2, "Nombre del Asegurado", "Persona Ficticia Editada", column_index=2),
            ),
            backup_dir=backup_dir,
        )

        assert backup_path.exists()
        backup_workbook = load_workbook(backup_path)
        saved_workbook = load_workbook(source)
        try:
            backup_sheet = backup_workbook[MAIN_SHEET_NAME]
            saved_sheet = saved_workbook[MAIN_SHEET_NAME]
            assert backup_sheet["A2"].value == "01-FICT-001"
            assert backup_sheet["B2"].value == "Persona Ficticia A"
            assert saved_sheet["A2"].value == "02-FICT-999"
            assert saved_sheet["B2"].value == "Persona Ficticia Editada"
            assert saved_sheet["E2"].value == "Cobertura Ficticia"
            assert saved_sheet["A3"].value == "01-FICT-002"
            assert "Hoja auxiliar" in saved_workbook.sheetnames
        finally:
            backup_workbook.close()
            saved_workbook.close()


def test_guardar_sobre_archivo_no_escribe_si_falla_respaldo(monkeypatch):
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "control_cartera.xlsx"
        backup_dir = temp_dir / "backups"
        build_control_cartera(source)

        def failing_copy(*args, **kwargs):
            raise OSError("fallo ficticio")

        monkeypatch.setattr("app.services.workbook_saver.shutil.copy2", failing_copy)

        with pytest.raises(WorkbookSaveError):
            save_control_cartera(
                source,
                (WorkbookCellUpdate(2, "Nº Póliza", "02-FICT-999", column_index=1),),
                backup_dir=backup_dir,
            )

        workbook = load_workbook(source)
        try:
            assert workbook[MAIN_SHEET_NAME]["A2"].value == "01-FICT-001"
        finally:
            workbook.close()
