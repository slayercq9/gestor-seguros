import json
import shutil
import subprocess
import sys
import uuid
from contextlib import contextmanager
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.core.exceptions import WorkbookModernizationError
from app.services.workbook_modernizer import REPORT_JSON_NAME, REPORT_MD_NAME, modernize_workbook


@contextmanager
def workspace_tempdir():
    base_dir = Path("data/output")
    base_dir.mkdir(parents=True, exist_ok=True)
    path = base_dir / f"tmp-modernizer-{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def build_fake_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cartera"
    sheet.append(
        [
            "Nombre Cliente",
            "Identificacion",
            "Numero de Poliza",
            "Vigencia",
            "Dia",
            "Mes",
            "Ano",
            "Detalle",
            "Numero de Placa / Finca",
        ]
    )
    sheet.append(["Registro Ficticio A", "ID-FICT-001", "POL-FICT-001", "D.M.", 1, 5, 2026, "Nota ficticia", "PLACA1"])
    sheet.append(["Registro Ficticio B", "ID-FICT-002", "POL-FICT-002", "Mensual", 15, 8, 2026, "Otra nota", "PLACA2"])
    sheet.append(["Registro Ficticio C", "ID-FICT-003", "POL-FICT-003", "Trimestral", 20, 10, 2026, "", ""])
    sheet.cell(row=30, column=1).fill = PatternFill("solid", fgColor="FFFFFF")
    workbook.save(path)


def build_fake_workbook_with_legacy_gs(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cartera"
    sheet.append(["Cliente", "Poliza", "Vigencia", "GS_ES_DM", "GS_REQUIERE_REVISION"])
    sheet.append(["Registro Ficticio A", "POL-FICT-001", "D.M.", "si", "no"])
    workbook.save(path)


def test_moderniza_copia_sin_alterar_fuente():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "base_ficticia.xlsx"
        output_dir = temp_dir / "salida"
        build_fake_workbook(source)
        original_bytes = source.read_bytes()

        result = modernize_workbook(source, output_dir)

        assert source.read_bytes() == original_bytes
        assert result.output_workbook.parent == output_dir.resolve()
        assert result.output_workbook.name.startswith("base_ficticia_modernizado_")
        assert result.output_workbook.name.endswith(".xlsx")
        assert result.output_workbook.exists()
        assert result.markdown_report.exists()
        assert result.json_report.exists()
        assert not (output_dir / "control_revision.csv").exists()


def test_no_agrega_columnas_gs_ni_renombra_encabezados_originales():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "base_ficticia.xlsx"
        output_dir = temp_dir / "salida"
        build_fake_workbook(source)

        result = modernize_workbook(source, output_dir)
        workbook = load_workbook(result.output_workbook)
        try:
            sheet = workbook["Cartera"]
            headers = [cell.value for cell in sheet[1] if cell.value]
            assert headers == [
                "Nombre Cliente",
                "Identificacion",
                "Numero de Poliza",
                "Vigencia",
                "Dia",
                "Mes",
                "Ano",
                "Detalle",
                "Numero de Placa / Finca",
            ]
            assert not any(str(header).startswith("GS_") for header in headers)
            assert result.useful_rows == 3
            assert result.rows_skipped >= 1
        finally:
            workbook.close()


def test_elimina_columnas_gs_heredadas_solo_en_la_copia():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "base_con_gs_heredadas.xlsx"
        output_dir = temp_dir / "salida"
        build_fake_workbook_with_legacy_gs(source)
        original_bytes = source.read_bytes()

        result = modernize_workbook(source, output_dir)
        workbook = load_workbook(result.output_workbook)
        try:
            sheet = workbook["Cartera"]
            headers = [cell.value for cell in sheet[1] if cell.value]
            assert headers == ["Cliente", "Poliza", "Vigencia"]
            assert not any(str(header).startswith("GS_") for header in headers)
        finally:
            workbook.close()

        assert source.read_bytes() == original_bytes
        data = json.loads((output_dir / REPORT_JSON_NAME).read_text(encoding="utf-8"))
        assert data["legacy_auxiliary_columns_removed"] == 2


def test_reportes_no_exponen_datos_sensibles_ficticios():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "base_ficticia.xlsx"
        output_dir = temp_dir / "salida"
        build_fake_workbook(source)

        modernize_workbook(source, output_dir)
        combined = (output_dir / REPORT_MD_NAME).read_text(encoding="utf-8")
        combined += (output_dir / REPORT_JSON_NAME).read_text(encoding="utf-8")

        for token in ["Registro Ficticio A", "ID-FICT-001", "POL-FICT-001", "PLACA1", "Nota ficticia"]:
            assert token not in combined

        data = json.loads((output_dir / REPORT_JSON_NAME).read_text(encoding="utf-8"))
        assert data["useful_rows"] == 3
        assert data["columns_added"] == []
        assert data["privacy"]["contains_row_samples"] is False
        assert data["privacy"]["contains_sensitive_values"] is False


def test_formato_visual_basico_y_hoja_control():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "base_ficticia.xlsx"
        output_dir = temp_dir / "salida"
        build_fake_workbook(source)

        result = modernize_workbook(source, output_dir)
        workbook = load_workbook(result.output_workbook)
        try:
            sheet = workbook["Cartera"]
            assert sheet.freeze_panes == "A2"
            assert sheet.auto_filter.ref is not None
            assert "CONTROL_MODERNIZACION" in workbook.sheetnames
            control = workbook["CONTROL_MODERNIZACION"]
            assert control["A10"].value == "columnas_auxiliares_agregadas"
            assert control["B10"].value == 0
        finally:
            workbook.close()


def test_script_falla_de_forma_controlada_si_no_existe_entrada():
    with workspace_tempdir() as temp_dir:
        completed = subprocess.run(
            [
                sys.executable,
                "scripts/modernizar_workbook_local.py",
                str(temp_dir / "no_existe.xlsx"),
                str(temp_dir / "salida"),
            ],
            check=False,
            capture_output=True,
            text=True,
        )

        assert completed.returncode == 1
        assert "Error de modernizacion" in completed.stdout


def test_rechaza_salida_igual_al_archivo_original():
    with workspace_tempdir() as temp_dir:
        source = temp_dir / "base_ficticia.xlsx"
        build_fake_workbook(source)

        try:
            modernize_workbook(source, source)
        except WorkbookModernizationError as exc:
            assert "archivo original" in str(exc)
        else:
            raise AssertionError("Se esperaba WorkbookModernizationError")
